import json
import os
from datetime import datetime
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── CONFIG ──────────────────────────────────────────────
DATA_DIR = "data"
INPUT_FILE = os.path.join(DATA_DIR, "dce_monitors_raw.json")
OUTPUT_FILE = os.path.join(DATA_DIR, "dce_monitor_audit.xlsx")
L1_SNOW_HANDLE = "appopsgccdce"
L2_SNOW_HANDLE = "appl2"
# ────────────────────────────────────────────────────────

os.makedirs(DATA_DIR, exist_ok=True)

# ── HELPERS ─────────────────────────────────────────────


def get_asset_id(tags):
    for t in tags:
        if t.startswith("tr_application-asset-insight-id:"):
            return t.split(":", 1)[1]
    return "MISSING"


def get_snow_info(notifications):
    snow_handles = [n["handle"] for n in notifications if "servicenow" in n.get("handle", "").lower()]
    if not snow_handles:
        return "None", "None", "Gap - No SNOW"
    handle_str = ", ".join(snow_handles)
    if any(L1_SNOW_HANDLE in h for h in snow_handles):
        return handle_str, "L1 GCC DCE", "Correct"
    elif any(L2_SNOW_HANDLE in h for h in snow_handles):
        return handle_str, "L2 Direct", "Review Needed"
    else:
        return handle_str, "Other", "Investigate"


def convert_ts(ts):
    if not ts:
        return "N/A", "N/A"
    dt = datetime.fromtimestamp(ts)
    days_old = (datetime.now() - dt).days
    return dt.strftime("%Y-%m-%d"), days_old


def classify(monitor):
    status = monitor.get("status", "")
    tags = monitor.get("tags", [])
    modified = monitor.get("modified")
    muted = monitor.get("muted_until_ts")
    notifications = monitor.get("notifications", [])

    snow_handles = [n["handle"] for n in notifications if "servicenow" in n.get("handle", "").lower()]
    has_asset_id = any(t.startswith("tr_application-asset-insight-id:") for t in tags)
    is_managed = any(t == "managed-by:terraform" for t in tags)

    days_old = 0
    if modified:
        days_old = (datetime.now() - datetime.fromtimestamp(modified)).days

    if is_managed:
        return "Already Managed"
    if status == "No Data" and days_old > 365:
        return "Retire"
    if status == "No Data":
        return "Investigate"
    if not has_asset_id:
        return "Fix then Migrate"
    if not snow_handles:
        return "Fix then Migrate"
    if any(L2_SNOW_HANDLE in h for h in snow_handles) and not any(L1_SNOW_HANDLE in h for h in snow_handles):
        return "Fix then Migrate"
    if muted:
        return "Investigate"
    return "Migrate"


# ── STYLING ─────────────────────────────────────────────

COLORS = {
    "header_bg": "1F3864",
    "header_font": "FFFFFF",
    "Migrate": "C6EFCE",
    "Retire": "FFC7CE",
    "Fix then Migrate": "FFEB9C",
    "Investigate": "FFEB9C",
    "Already Managed": "BDD7EE",
    "Correct": "C6EFCE",
    "Review Needed": "FFEB9C",
    "Investigate_snow": "FFEB9C",
    "Gap - No SNOW": "FFC7CE",
    "row_alt": "F2F2F2",
}


def hdr_style(cell):
    cell.font = Font(bold=True, color=COLORS["header_font"], name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def color_cell(cell, hex_color):
    cell.fill = PatternFill("solid", fgColor=hex_color)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


# ── MAIN ────────────────────────────────────────────────


def build_audit(monitors):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Monitor Audit ───────────────────────────
    ws = wb.active
    ws.title = "Monitor Audit"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 35

    headers = [
        "Monitor ID",
        "Name",
        "Type",
        "Status",
        "Asset ID",
        "Has Asset ID",
        "Creator",
        "Created Date",
        "Last Modified",
        "Days Since Modified",
        "Muted",
        "SNOW Handle",
        "SNOW Queue",
        "SNOW Routing Flag",
        "Tags",
        "Recommendation",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        hdr_style(cell)

    for row_idx, m in enumerate(monitors, 2):
        tags = m.get("tags", [])
        notifications = m.get("notifications", [])
        modified_ts = m.get("modified")
        created_ts = m.get("created")
        muted = m.get("muted_until_ts")

        modified_date, days_old = convert_ts(modified_ts)
        created_date, _ = convert_ts(created_ts)
        asset_id = get_asset_id(tags)
        snow_handle, snow_queue, snow_flag = get_snow_info(notifications)
        recommendation = classify(m)

        # Alt row background
        bg = COLORS["row_alt"] if row_idx % 2 == 0 else "FFFFFF"

        row_data = [
            m.get("id"),
            m.get("name"),
            m.get("type"),
            m.get("status", "Unknown"),
            asset_id,
            "Yes" if asset_id != "MISSING" else "No",
            m.get("creator", {}).get("handle", "Unknown"),
            created_date,
            modified_date,
            days_old,
            "Yes" if muted else "No",
            snow_handle,
            snow_queue,
            snow_flag,
            ", ".join(tags),
            recommendation,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = thin_border()
            cell.fill = PatternFill("solid", fgColor=bg)

        # Colour Status column (D = col 4)
        status_colors = {
            "OK": "C6EFCE",
            "No Data": "FFC7CE",
            "Alert": "FF0000",
            "Warn": "FFEB9C",
        }
        status_val = m.get("status", "")
        if status_val in status_colors:
            color_cell(ws.cell(row=row_idx, column=4), status_colors[status_val])

        # Colour SNOW Routing Flag (col 14)
        snow_flag_colors = {
            "Correct": "C6EFCE",
            "Review Needed": "FFEB9C",
            "Investigate": "FFEB9C",
            "Gap - No SNOW": "FFC7CE",
        }
        if snow_flag in snow_flag_colors:
            color_cell(ws.cell(row=row_idx, column=14), snow_flag_colors[snow_flag])

        # Colour Recommendation (col 16)
        rec_colors = {
            "Migrate": "C6EFCE",
            "Retire": "FFC7CE",
            "Fix then Migrate": "FFEB9C",
            "Investigate": "FFEB9C",
            "Already Managed": "BDD7EE",
        }
        if recommendation in rec_colors:
            color_cell(ws.cell(row=row_idx, column=16), rec_colors[recommendation])

    # Column widths
    col_widths = [12, 55, 18, 12, 12, 12, 35, 13, 13, 10, 8, 40, 15, 18, 50, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Sheet 2: Summary ────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 15

    rec_counts = Counter(classify(m) for m in monitors)
    status_counts = Counter(m.get("status", "Unknown") for m in monitors)
    snow_counts = Counter(get_snow_info(m.get("notifications", []))[1] for m in monitors)
    no_asset = sum(1 for m in monitors if get_asset_id(m.get("tags", [])) == "MISSING")
    muted_count = sum(1 for m in monitors if m.get("muted_until_ts"))

    sections = [
        ("DCE L1 Monitor Audit — Summary", None),
        (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", None),
        ("", None),
        ("TOTAL MONITORS IN SCOPE", len(monitors)),
        ("", None),
        ("── By Status ──", None),
        *[(k, v) for k, v in sorted(status_counts.items())],
        ("", None),
        ("── By Recommendation ──", None),
        *[(k, v) for k, v in sorted(rec_counts.items())],
        ("", None),
        ("── ServiceNow Routing ──", None),
        *[(k, v) for k, v in sorted(snow_counts.items())],
        ("", None),
        ("── Tag Gaps ──", None),
        ("Missing Asset ID tag", no_asset),
        ("Muted monitors", muted_count),
    ]

    for r_idx, (label, value) in enumerate(sections, 1):
        a = ws2.cell(row=r_idx, column=1, value=label)
        a.font = Font(name="Arial", size=10)
        if value is not None:
            b = ws2.cell(row=r_idx, column=2, value=value)
            b.font = Font(name="Arial", size=10)
            b.alignment = Alignment(horizontal="center")
        if label.startswith("──") or label == "TOTAL MONITORS IN SCOPE":
            a.font = Font(name="Arial", size=10, bold=True)
        if r_idx == 1:
            a.font = Font(name="Arial", size=13, bold=True, color=COLORS["header_bg"])

    wb.save(OUTPUT_FILE)
    return rec_counts, status_counts, snow_counts, no_asset, muted_count


# ── RUN ─────────────────────────────────────────────────

if __name__ == "__main__":
    if not os.path.exists(INPUT_FILE):
        print(f"Error: {INPUT_FILE} not found. Run extract_dce_monitors.py first.")
        exit(1)

    with open(INPUT_FILE, "r", encoding="utf-8") as f:
        monitors = json.load(f)

    print(f"Processing {len(monitors)} monitors...")
    rec_counts, status_counts, snow_counts, no_asset, muted_count = build_audit(monitors)

    print(f"\n✅ Audit saved to {OUTPUT_FILE}")
    print(f"\n{'=' * 40}")
    print("SUMMARY")
    print(f"{'=' * 40}")
    print(f"Total monitors: {len(monitors)}")
    print("\nBy Status:")
    for k, v in sorted(status_counts.items()):
        print(f"  {k}: {v}")
    print("\nBy Recommendation:")
    for k, v in sorted(rec_counts.items()):
        print(f"  {k}: {v}")
    print("\nServiceNow Routing:")
    for k, v in sorted(snow_counts.items()):
        print(f"  {k}: {v}")
    print(f"\nMissing Asset ID tag: {no_asset}")
    print(f"Muted monitors: {muted_count}")
